# teaor_profiles.py
# ------------------------------
# Load TEÁOR -> bin probability profiles from a custom Excel sheet where
# TEÁOR code is in column K, bin label in column M, and probability/percent in column J.
#
# This loader is robust:
# - tries pandas first (fast),
# - falls back to openpyxl if pandas isn't installed,
# - normalizes per TEÁOR (sum to 1),
# - ignores missing/invalid rows,
# - sorts bins numerically by lower bound.

from __future__ import annotations

from typing import Dict, Tuple, List, Optional


VALID_BINS = {"1-4", "5-9", "10-19", "20-49", "50-249", "250-1000"}


def _bin_lower(b: str) -> int:
    # "50-249" -> 50
    return int(b.split("-")[0])


def _normalize_bin_probs(bin_probs: Dict[str, float]) -> Dict[str, float]:
    total = sum(bin_probs.values())
    if total <= 0:
        return {}
    return {k: v / total for k, v in bin_probs.items()}


def _coerce_teaor(v) -> Optional[str]:
    """
    Accepts values like:
      - 1, 1.0, "1", "01"
      - "teaor 1", "TEAOR 43", "teaor: 85", etc.
    Returns numeric TEÁOR code as string without leading zeros: "1".."96".
    """
    if v is None:
        return None

    s = str(v).strip()
    if s == "":
        return None

    # Normalize
    s_norm = s.lower().replace(":", " ").replace("_", " ")
    # If it contains the word "teaor", try to extract the last integer-like token
    if "teaor" in s_norm:
        parts = s_norm.split()
        # find any numeric token; prefer the last one
        nums = []
        for p in parts:
            p = p.replace(",", ".")
            try:
                f = float(p)
                if f.is_integer():
                    nums.append(str(int(f)))
            except Exception:
                continue
        if nums:
            return nums[-1]  # e.g. "teaor 43" -> "43"

    # handle plain numeric strings (including floats like "43.0")
    try:
        f = float(s.replace(",", "."))
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass

    # As a last resort, if string ends with digits, extract them
    digits = ""
    for ch in reversed(s):
        if ch.isdigit():
            digits = ch + digits
        else:
            break
    return digits if digits else None


def _coerce_bin(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None

    # Normalize dash and spaces: "5 – 9" / "5- 9" / "5 -9" -> "5-9"
    s = s.replace("–", "-")
    s = s.replace("—", "-")
    s = s.replace(" ", "")

    # now it should look like "5-9"
    if s in VALID_BINS:
        return s
    return None



def _coerce_prob(v) -> Optional[float]:
    """
    Column J can be percent-like numbers (e.g. 98.54) or already 0..1.
    We'll interpret:
      - if > 1.0 => assume percent and divide by 100
      - else => assume probability already
    """
    if v is None:
        return None
    try:
        if isinstance(v, str):
            v = v.strip().replace(",", ".")
        x = float(v)
    except Exception:
        return None

    if x < 0:
        return None
    if x > 1.0:
        x = x / 100.0
    return x


def load_teaor_bin_profiles(
    excel_path: str,
    sheet_name: Optional[str] = None,
    start_row: int = 1,
) -> Dict[str, Dict[str, float]]:
    """
    Reads an Excel file where:
      - J column: probability/percent
      - K column: teaor code
      - M column: bin label

    start_row: 1-based row index to begin reading (set to 1 unless you have headers above).
    sheet_name: if None, uses the first sheet.
    """
    # Try pandas first
    try:
        import pandas as pd  # type: ignore

        df = pd.read_excel(
            excel_path,
            sheet_name=sheet_name if sheet_name is not None else 0,
            usecols="J,K,M",
            header=None,
            skiprows=max(0, start_row - 1),
            engine="openpyxl",
        )

        # Forward-fill TEÁOR column (col 1) because only first row has the label
        df.iloc[:, 1] = df.iloc[:, 1].ffill()

        # Columns: 0=J(prob), 1=K(teaor), 2=M(bin)
        profiles: Dict[str, Dict[str, float]] = {}

        for _, row in df.iterrows():
            prob = _coerce_prob(row.iloc[0])
            tea = _coerce_teaor(row.iloc[1])
            b = _coerce_bin(row.iloc[2])

            if prob is None or tea is None or b is None:
                continue
            if prob == 0:
                continue

            profiles.setdefault(tea, {})
            profiles[tea][b] = profiles[tea].get(b, 0.0) + prob

        # normalize each TEÁOR
        out: Dict[str, Dict[str, float]] = {}
        for tea, bp in profiles.items():
            nbp = _normalize_bin_probs(bp)
            if nbp:
                # sort bins nicely (optional: keep dict but in sorted insertion order)
                for bname in sorted(nbp.keys(), key=_bin_lower):
                    out.setdefault(tea, {})[bname] = nbp[bname]
        return out

    except Exception:
        # Fallback to openpyxl
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name] if sheet_name is not None else wb.worksheets[0]

        # Convert Excel col letters to indices (1-based)
        colJ = 10
        colK = 11
        colM = 13

        profiles: Dict[str, Dict[str, float]] = {}

        for r in range(start_row, ws.max_row + 1):
            prob = _coerce_prob(ws.cell(row=r, column=colJ).value)
            tea = _coerce_teaor(ws.cell(row=r, column=colK).value)
            b = _coerce_bin(ws.cell(row=r, column=colM).value)

            if prob is None or tea is None or b is None:
                continue
            if prob == 0:
                continue

            profiles.setdefault(tea, {})
            profiles[tea][b] = profiles[tea].get(b, 0.0) + prob

        out: Dict[str, Dict[str, float]] = {}
        for tea, bp in profiles.items():
            nbp = _normalize_bin_probs(bp)
            if nbp:
                for bname in sorted(nbp.keys(), key=_bin_lower):
                    out.setdefault(tea, {})[bname] = nbp[bname]
        return out


def pretty_print_profile(profiles: Dict[str, Dict[str, float]], teaor: str, max_bins: int = 10) -> None:
    """
    Debug helper.
    """
    p = profiles.get(teaor)
    if not p:
        print(f"No profile for TEÁOR={teaor}")
        return
    items = list(p.items())[:max_bins]
    print(f"TEÁOR {teaor}:")
    for b, v in items:
        print(f"  {b}: {v:.6f}")
    print(f"  sum: {sum(p.values()):.6f}")
